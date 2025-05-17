# src/report_generator.py

import pandas as pd
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from src.utils import flatten_list_to_string, ensure_output_dir, get_timestamped_filename

# At the top of export_to_excel()
summary = parsed_data.get("summary", {})
summary_lines = {
    "threat_level": summary.get("threat_level", "N/A"),
    "recommendation": summary.get("recommendation", "N/A"),
    "reason": summary.get("reason", "N/A")
}

# Combine summary + rest of parsed data
combined_data = {**summary_lines, **parsed_data}
formatted_data = {
    key: flatten_list_to_string(value)
    for key, value in combined_data.items()
}


def export_to_excel(parsed_data, output_dir="output"):
    ensure_output_dir(output_dir)

    # Flatten list values for Excel
    formatted_data = {
        key: flatten_list_to_string(value)
        for key, value in parsed_data.items()
    }

    # Create DataFrame
    df = pd.DataFrame.from_dict(formatted_data, orient='index', columns=['Value'])

    # Save to Excel
    filename = get_timestamped_filename(extension="xlsx")
    filepath = os.path.join(output_dir, filename)
    df.to_excel(filepath, engine='openpyxl')

    # Load workbook and worksheet for formatting
    wb = load_workbook(filepath)
    ws = wb.active

    # Define styles
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply formatting to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment
            cell.border = border_style
            if cell.row == 1:
                cell.font = header_font

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 100)

    wb.save(filepath)
    print(f"[✓] Excel report saved to {filepath}")


